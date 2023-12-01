from tkinter import *
from tkinter import ttk
import sys,mysql.connector
from openpyxl import Workbook
from tkinter import messagebox
import smtplib
from email.mime.text import MIMEText
import secrets
import string
#Hàm connect db.
add_window = None

def connection():
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="camly",
            database="qlydiemsv"
        )
        return conn
    except mysql.connector.Error as e:
        print("Error connecting to the database:", e)
        return None

# Sinh viên Start #########################################################
def view_sinhvien():
    # Sử dụng biến global.
    global tree
    global search_sinhvien_entry
    # Xóa nội dung cũ trong Text widget

    # Tạo input search và button search và clear search
    search_sinhvien_entry = Entry(root, width=40)
    search_sinhvien_entry.grid(row=6, column=0, padx=0, pady=5)

    search_sinhvien_button = Button(root, text="Tìm kiếm", command=search_sinhvien, width=20)
    search_sinhvien_button.grid(row=7, column=0, padx=0, pady=5)

    add_sinhvien_button = Button(root, text="Thêm sinh viên", command=add_sinhvien_window, width=20)
    add_sinhvien_button.grid(row=12, column=0, padx=0, pady=5)

    conn = connection()
    cur = conn.cursor()
    cur.execute("select * from sinhvien")
    data = cur.fetchall()
    conn.close()

    # Tạo một Treeview widget
    tree = ttk.Treeview(root, columns=(1, 2, 3, 4, 5, 6, 7), show="headings", height=20)

    # Đặt tên các cột
    tree.heading(1, text="Mã SV")
    tree.heading(2, text="Họ và tên SV")
    tree.heading(3, text="Ngày sinh")
    tree.heading(4, text="Giới tính")
    tree.heading(5, text="Dân tộc")
    tree.heading(6, text="Nơi sinh")
    tree.heading(7, text="Mã lớp")
    
    # Đặt lại chiều rộng của mã SV, Mã lớp, số tín chỉ.
    tree.column(1, width=60)
    tree.column(6, width=150)
    

    # Hiển thị dữ liệu trong Treeview
    for row in data:
        tree.insert("", "end", values=row)

    # Thêm Treeview vào cửa sổ chính
    tree.grid(row=0, column=1, pady=0, padx=10,rowspan=26)

    # Tạo và thiết lập Scrollbar
    scrollbar = ttk.Scrollbar(root, orient="vertical", command=tree.yview)
    scrollbar.grid(row=0, column=2, sticky="ns",rowspan=26)
    tree.configure(yscrollcommand=scrollbar.set)

    def on_item_click(event):
        # Lấy thông tin của dòng được chọn
        selected_item = tree.selection()[0]
        values = tree.item(selected_item, 'values')

        # Hiển thị cửa sổ chi tiết điểm học phần
        view_sv_details(values)

    # Gán hàm on_item_click khi click vào dòng
    tree.bind('<ButtonRelease-1>', on_item_click)

def search_sinhvien():

    global tree 

    search_query = search_sinhvien_entry.get().lower()
    conn = connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM sinhvien WHERE LOWER(hoten_sv) LIKE %s OR ma_sv LIKE %s OR ma_lop LIKE %s OR LOWER(gioi_tinh) LIKE %s", ('%' + search_query + '%', '%' + search_query + '%', '%' + search_query + '%', '%' + search_query + '%'))
    data = cur.fetchall()
    conn.close()

    # Clear the Treeview
    for item in tree.get_children():
        tree.delete(item)

    # Display the filtered data in the Treeview
    for row in data:
        tree.insert("", "end", values=row)

# Hiển thị form thêm mới sinh viên
def add_sinhvien_window():
    add_window = Toplevel(root)
    add_window.title("Thêm Sinh Viên")

    Label(add_window, text="Mã SV:").grid(row=0, column=0, padx=10, pady=5)
    ma_sv_entry = Entry(add_window, width=40)
    ma_sv_entry.grid(row=0, column=1, padx=10, pady=5)

    Label(add_window, text="Họ và tên SV:").grid(row=1, column=0, padx=10, pady=5)
    hoten_sv_entry = Entry(add_window, width=40)
    hoten_sv_entry.grid(row=1, column=1, padx=10, pady=5)

    Label(add_window, text="Ngày sinh (YYYY-MM-DD):").grid(row=2, column=0, padx=10, pady=5)
    ngay_sinh_entry = Entry(add_window, width=40)
    ngay_sinh_entry.grid(row=2, column=1, padx=10, pady=5)

    Label(add_window, text="Giới tính (Nam/Nữ):").grid(row=3, column=0, padx=10, pady=5)
    gioi_tinh_entry = Entry(add_window, width=40)
    gioi_tinh_entry.grid(row=3, column=1, padx=10, pady=5)

    Label(add_window, text="Dân tộc:").grid(row=4, column=0, padx=10, pady=5)
    dan_toc_entry = Entry(add_window, width=40)
    dan_toc_entry.grid(row=4, column=1, padx=10, pady=5)

    Label(add_window, text="Nơi sinh:").grid(row=5, column=0, padx=10, pady=5)
    noi_sinh_entry = Entry(add_window, width=40)
    noi_sinh_entry.grid(row=5, column=1, padx=10, pady=5)

    # Combobox get mã lớp
    Label(add_window, text="Mã lớp:").grid(row=6, column=0, padx=10, pady=5)
    ma_lop_values = get_ma_lop_values()
    ma_lop_combobox = ttk.Combobox(add_window, values=ma_lop_values, width=37)
    ma_lop_combobox.grid(row=6, column=1, padx=10, pady=5)

    add_button = Button(add_window, text="Thêm Sinh Viên", command=lambda: add_sinhvien(
        ma_sv_entry.get(), hoten_sv_entry.get(), ngay_sinh_entry.get(),
        gioi_tinh_entry.get(), dan_toc_entry.get(), noi_sinh_entry.get(), ma_lop_combobox.get()
    ), width=20)
    add_button.grid(row=7, column=1, padx=10, pady=10)
   
#thêm sinh viên
def add_sinhvien(ma_sv, hoten_sv, ngay_sinh, gioi_tinh, dan_toc, noi_sinh, ma_lop):
    conn = connection()
    cur = conn.cursor()
    try:
        # Kiểm tra xem mã sinh viên đã tồn tại hay chưa
        cur.execute("SELECT ma_sv FROM sinhvien WHERE ma_sv = %s", (ma_sv,))
        existing_sv = cur.fetchone()

        if existing_sv:
            conn.close()
            messagebox.showerror("Lỗi", "Mã sinh viên đã tồn tại")
        else:
            # Thêm sinh viên nếu mã sinh viên chưa tồn tại
            cur.execute("""
                INSERT INTO sinhvien (ma_sv, hoten_sv, ngay_sinh, gioi_tinh, dan_toc, noi_sinh, ma_lop) 
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (ma_sv, hoten_sv, ngay_sinh, gioi_tinh, dan_toc, noi_sinh, ma_lop))
            conn.commit()
            conn.close()
            messagebox.showinfo("Thông báo", "Thêm sinh viên thành công")
            search_sinhvien()
    except mysql.connector.Error as e:
        print("Error adding data to database:", e)
        conn.rollback()
        conn.close()

def view_sv_details(student_info):
    # Tạo một cửa sổ mới
    details_window = Toplevel(root)
    details_window.title("SV:" + student_info[1]) 

    # Hiển thị thông tin chi tiết của dòng được chọn
    Label(details_window, text="Mã SV:").grid(row=0, column=0, padx=10, pady=5)
    Label(details_window, text=student_info[0]).grid(row=0, column=1, padx=10, pady=5)

    Label(details_window, text="Họ và tên SV:").grid(row=1, column=0, padx=10, pady=5)
    # Label(details_window, text=student_info[1]).grid(row=1, column=1, padx=10, pady=5)
    name_entry = Entry(details_window)
    name_entry.grid(row=1, column=1, padx=10, pady=5)
    name_entry.insert(0, student_info[1]) 
    
    Label(details_window, text="Năm Sinh:").grid(row=2, column=0, padx=10, pady=5)
    # Label(details_window, text=student_info[2]).grid(row=2, column=1, padx=10, pady=5)
    birthday_entry = Entry(details_window)
    birthday_entry.grid(row=2, column=1, padx=10, pady=5)
    birthday_entry.insert(0, student_info[2]) 
    Label(details_window, text="Giới Tính:").grid(row=3, column=0, padx=10, pady=5)
    # Label(details_window, text=student_info[3]).grid(row=3, column=1, padx=10, pady=5)
    gt_entry = Entry(details_window)
    gt_entry.grid(row=3, column=1, padx=10, pady=5)
    gt_entry.insert(0, student_info[3]) 
    Label(details_window, text="Dân Tộc:").grid(row=4, column=0, padx=10, pady=5)
    # Label(details_window, text=student_info[4]).grid(row=4, column=1, padx=10, pady=5)
    dt_entry = Entry(details_window)
    dt_entry.grid(row=4, column=1, padx=10, pady=5)
    dt_entry.insert(0, student_info[4]) 
    Label(details_window, text="Địa Chỉ:").grid(row=5, column=0, padx=10, pady=5)
    addr_entry = Entry(details_window)
    addr_entry.grid(row=5, column=1, padx=10, pady=5)
    addr_entry.insert(0, student_info[5]) 
    delete_button = Button(details_window, text="Xoá", command=lambda: delete_sinhvien(student_info[0]), fg="white", bg="red",width=20)
    delete_button.grid(row=8, column=0, padx=10, pady=5)

    # #cập nhật tt sinh vien
    def update_sv():
        name = name_entry.get()
        birthday = birthday_entry.get()
        gt = gt_entry.get()
        dt = dt_entry.get()
        addr = addr_entry.get()
        lop = class_entry.get()
        total = total_entry.get()
    
        conn = connection()
        cur = conn.cursor()
        cur.execute("UPDATE sinhvien SET hoten_sv=%s, ngay_sinh=%s, gioi_tinh=%s, dan_toc=%s, noi_sinh=%s, ma_lop=%s WHERE ma_sv=%s",
            (name, birthday, gt, dt, addr, lop, student_info[0]))
     #print("da va0" + new_final)
        messagebox.showinfo("Success", "Thông tin sinh viên đã được cập nhật!")
            # Cập nhật lại dữ liệu trong Treeview
         # view_diemhocphan()
        
        conn.commit()
        conn.close()
            # Đóng cửa sổ chi tiết sau khi cập nhật
        details_window.destroy()
        view_sinhvien()
    #xoá sinh viên 
    def delete_sinhvien(ma_sv):
        conn = connection()
        cur = conn.cursor()
        try:
            # Kiểm tra xem sinh viên tồn tại hay không trước khi xoá
            cur.execute("SELECT ma_sv FROM sinhvien WHERE ma_sv = %s", (ma_sv,))
            existing_sv = cur.fetchone()

            if existing_sv:
                # Nếu sinh viên tồn tại, thực hiện xoá
                cur.execute("DELETE FROM sinhvien WHERE ma_sv = %s", (ma_sv,))
                conn.commit()
                conn.close()
                messagebox.showinfo("Thông báo", "Xoá sinh viên thành công")
                details_window.destroy()
                search_sinhvien()
            else:
                # Nếu sinh viên không tồn tại, hiển thị thông báo lỗi
                conn.close()
                messagebox.showerror("Lỗi", "Không tìm thấy sinh viên có mã này")
        except mysql.connector.Error as e:
            print("Error deleting data from database:", e)
            conn.rollback()
            conn.close()
    update_button = Button(details_window, text="Cập nhật", command=update_sv, fg="black", bg="yellow", width=20)
    update_button.grid(row=8, columnspan=2, column=1, padx=10, pady=5)
    # Combobox get mã lớp
    Label(details_window, text="Mã lớp:").grid(row=6, column=0, padx=10, pady=5)
    ma_lop_values = get_ma_lop_values()
    class_entry = ttk.Combobox(details_window, values=ma_lop_values, width=17)
    class_entry.set(student_info[6])
    class_entry.grid(row=6, column=1, padx=10, pady=5)
    total_credit_hours = total_credit_hours_of_student(student_info[0])
    Label(details_window, text="Tổng số tín chỉ đã học :").grid(row=7, column=0, padx=10, pady=5)
    total_entry = Entry(details_window)
    Label(details_window, text=total_credit_hours).grid(row=7, column=1, padx=10, pady=5)
    total_entry.insert(0, student_info[7])  # Hiển thị giá trị hiện tại
    
# Sinh viên End #########################################################

# Điểm học phần Start #########################################################
def view_diemhocphan():
    # Sử dụng biến global.
    global tree
    global search_diemhocphan_entry
    # Xóa nội dung cũ trong Text widget

    # Tạo input search và button search và clear search
    search_diemhocphan_entry = Entry(root, width=40)
    search_diemhocphan_entry.grid(row=6, column=0, padx=0, pady=5)

    search_diemhocphan_button = Button(root, text="Tìm kiếm", command=search_diemhocphan, width=20)
    search_diemhocphan_button.grid(row=7, column=0, padx=0, pady=5)

    add_diemhocphan_button = Button(root, text="Thêm điểm học phần", command=add_diemhocphan_window, width=20)
    add_diemhocphan_button.grid(row=12, column=0, padx=0, pady=5)

    conn = connection()
    cur = conn.cursor()
    cur.execute("select diemhocphan.ma_sv,hoten_sv,ma_lop,diemhocphan.ma_mon,ten_mon,sotinchi,diem_giua_ky,diem_thi_hp from diemhocphan inner join sinhvien on diemhocphan.ma_sv=sinhvien.ma_sv inner join monhocphan on monhocphan.ma_mon=diemhocphan.ma_mon")
    data = cur.fetchall()
    conn.close()

    # Tạo một Treeview widget
    tree = ttk.Treeview(root, columns=(1, 2, 3, 4, 5, 6, 7, 8), show="headings", height=20)

    # Đặt tên các cột
    tree.heading(1, text="Mã SV")
    tree.heading(2, text="Họ và tên SV")
    tree.heading(3, text="Mã lớp")
    tree.heading(4, text="Mã học phần")
    tree.heading(5, text="Tên học phần")
    tree.heading(6, text="Số tín chỉ")
    tree.heading(7, text="Điểm giữa kỳ")
    tree.heading(8, text="Điểm cuối kỳ")
    
    # Đặt lại chiều rộng của mã SV, Mã lớp, số tín chỉ.
    tree.column(1, width=60)
    tree.column(3, width=90)
    tree.column(6, width=60)
    

    # Hiển thị dữ liệu trong Treeview
    for row in data:
        tree.insert("", "end", values=row)

    # Thêm Treeview vào cửa sổ chính
    tree.grid(row=0, column=1, pady=0, padx=10,rowspan=26)

    # Tạo và thiết lập Scrollbar
    scrollbar = ttk.Scrollbar(root, orient="vertical", command=tree.yview)
    scrollbar.grid(row=0, column=2, sticky="ns",rowspan=26)
    tree.configure(yscrollcommand=scrollbar.set)

    def on_item_click(event):
        # Lấy thông tin của dòng được chọn
        selected_item = tree.selection()[0]
        values = tree.item(selected_item, 'values')

        # Hiển thị cửa sổ chi tiết điểm học phần
        # view_diemhocphan_details(values)
        view_Diemhp_details(values)
    # Gán hàm on_item_click khi click vào dòng
    tree.bind('<ButtonRelease-1>', on_item_click)

def view_Diemhp_details(student_info):
    # Tạo một cửa sổ mới
    details_window = Toplevel(root)
    details_window.title("SV:" + student_info[1]) 

    # Hiển thị thông tin chi tiết của dòng được chọn
    Label(details_window, text="Mã SV:").grid(row=0, column=0, padx=10, pady=5)
    Label(details_window, text=student_info[0]).grid(row=0, column=1, padx=10, pady=5)

    Label(details_window, text="Họ và tên SV:").grid(row=1, column=0, padx=10, pady=5)
    Label(details_window, text=student_info[1]).grid(row=1, column=1, padx=10, pady=5)

    Label(details_window, text="Mã lớp:").grid(row=2, column=0, padx=10, pady=5)
    Label(details_window, text=student_info[2]).grid(row=2, column=1, padx=10, pady=5)

    Label(details_window, text="Mã học phần:").grid(row=3, column=0, padx=10, pady=5)
    Label(details_window, text=student_info[3]).grid(row=3, column=1, padx=10, pady=5)

    Label(details_window, text="Tên học phần:").grid(row=4, column=0, padx=10, pady=5)
    Label(details_window, text=student_info[4]).grid(row=4, column=1, padx=10, pady=5)

    Label(details_window, text="Số tín chỉ:").grid(row=5, column=0, padx=10, pady=5)
    Label(details_window, text=student_info[5]).grid(row=5, column=1, padx=10, pady=5)

    # Sử dụng Entry để cho phép chỉnh sửa
    Label(details_window, text="Điểm giữa kỳ:").grid(row=6, column=0, padx=10, pady=5)
    midterm_entry = Entry(details_window)
    midterm_entry.grid(row=6, column=1, padx=10, pady=5)
    midterm_entry.insert(0, student_info[6])  # Hiển thị giá trị hiện tại

    Label(details_window, text="Điểm thi HP:").grid(row=7, column=0, padx=10, pady=5)
    final_entry = Entry(details_window)
    final_entry.grid(row=7, column=1, padx=10, pady=5)
    final_entry.insert(0, student_info[7])  # Hiển thị giá trị hiện tại

    # Hàm cập nhật điểm
    def update_grade():
        # Lấy giá trị mới từ các ô nhập liệu
        new_midterm = midterm_entry.get()
        new_final = final_entry.get()

        # Thực hiện cập nhật vào cơ sở dữ liệu
        conn = connection()
        cur = conn.cursor()
        cur.execute("UPDATE diemhocphan SET diem_giua_ky=%s, diem_thi_hp=%s WHERE ma_sv=%s AND ma_mon=%s", (new_midterm, new_final, student_info[0], student_info[3]))
        conn.commit()
        conn.close()

        # Cập nhật lại dữ liệu trong Treeview
        view_diemhocphan()

        # Đóng cửa sổ chi tiết sau khi cập nhật
        details_window.destroy()


    # Nút cập nhật điểm
    update_button = Button(details_window, text="Cập nhật điểm", command=update_grade, fg="black", bg="yellow", width=20)
    update_button.grid(row=8, column=1, padx=10, pady=10)


def search_diemhocphan():

    global tree 

    search_query = search_diemhocphan_entry.get().lower()
    conn = connection()
    cur = conn.cursor()
    cur.execute("SELECT diemhocphan.ma_sv,hoten_sv,ma_lop,diemhocphan.ma_mon,ten_mon,sotinchi,diem_giua_ky,diem_thi_hp FROM diemhocphan INNER JOIN sinhvien ON diemhocphan.ma_sv=sinhvien.ma_sv INNER JOIN monhocphan ON monhocphan.ma_mon=diemhocphan.ma_mon WHERE LOWER(hoten_sv) LIKE %s OR diemhocphan.ma_sv LIKE %s OR diemhocphan.ma_mon LIKE %s OR ma_lop LIKE %s OR LOWER(ten_mon) LIKE %s", ('%' + search_query + '%', '%' + search_query + '%', '%' + search_query + '%', '%' + search_query + '%', '%' + search_query + '%'))
    data = cur.fetchall()
    conn.close()

    # Clear the Treeview
    for item in tree.get_children():
        tree.delete(item)

    # Display the filtered data in the Treeview
    for row in data:
        tree.insert("", "end", values=row)

def add_diemhocphan_window():
    global add_window
    global ma_sv_combobox
    global ma_lop_combobox
    add_window = Toplevel(root)
    add_window.title("Thêm Điểm Học Phần")

    # Combobox get mã lớp
    Label(add_window, text="Mã lớp:").grid(row=0, column=0, padx=10, pady=5)
    ma_lop_values = get_ma_lop_values()
    ma_lop_combobox = ttk.Combobox(add_window, values=ma_lop_values, width=37)
    ma_lop_combobox.grid(row=0, column=1, padx=10, pady=5)

    ma_lop_combobox.bind("<<ComboboxSelected>>", update_sinhvien_combobox)

     # Combobox get mã sinh viên
    Label(add_window, text="Mã SV:").grid(row=1, column=0, padx=10, pady=5)
    ma_sv_combobox = ttk.Combobox(add_window, width=37)
    ma_sv_combobox.grid(row=1, column=1, padx=10, pady=5)

    Label(add_window, text="Mã Môn Học:").grid(row=2, column=0, padx=10, pady=5)
    ma_mon_values = get_ma_mon_values()
    ma_mon_combobox = ttk.Combobox(add_window, values=ma_mon_values, width=37)
    ma_mon_combobox.grid(row=2, column=1, padx=10, pady=5)

    Label(add_window, text="Điểm Giữa Kỳ:").grid(row=3, column=0, padx=10, pady=5)
    diem_giua_ky_entry = Entry(add_window, width=40)
    diem_giua_ky_entry.grid(row=3, column=1, padx=10, pady=5)

    Label(add_window, text="Điểm Thi HP:").grid(row=4, column=0, padx=10, pady=5)
    diem_thi_hp_entry = Entry(add_window, width=40)
    diem_thi_hp_entry.grid(row=4, column=1, padx=10, pady=5)

    # Thêm nút để thực hiện thêm điểm học phần
    add_button = Button(add_window, text="Thêm Điểm", command=lambda: add_diemhocphan(ma_sv_combobox.get(), ma_mon_combobox.get(), diem_giua_ky_entry.get(), diem_thi_hp_entry.get()), width=20)
    add_button.grid(row=5, column=1, padx=10, pady=10)


def add_diemhocphan(ma_sv, ma_mon, diem_giua_ky, diem_thi_hp):
    global add_window
    # xử lý chỉ lấy mã sv loại bỏ phần tên sv. vd: DTC12 {Hoàng C} tách DTC12 từ chuỗi này.
    masvSplit = ma_sv.split('{')[0].strip()
    conn = connection()
    cur = conn.cursor()
    try:
        # Try thêm dữ liệu
        cur.execute("""
            INSERT INTO diemhocphan (ma_sv, ma_mon, diem_giua_ky, diem_thi_hp) 
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE 
            diem_giua_ky = VALUES(diem_giua_ky), diem_thi_hp = VALUES(diem_thi_hp)
        """, (masvSplit, ma_mon, diem_giua_ky, diem_thi_hp))
        conn.commit()
        search_diemhocphan()  # Load lại dữ liệu
        add_window.destroy()  # Đóng cửa sổ
        if add_window:
            add_window.destroy()
            messagebox.showinfo("Success", "Cập nhật thành công!")
        conn.close()
    except mysql.connector.Error as e:
        print("Thêm thất bại", e)
        conn.rollback()
        conn.close()

# Điểm học phần End ####################################

# Học phần Start #########################################################
def view_hocphan():
    # Sử dụng biến global.
    global tree
    global search_hocphan_entry
    # Xóa nội dung cũ trong Text widget

    # Tạo input search và button search và clear search
    search_hocphan_entry = Entry(root, width=40)
    search_hocphan_entry.grid(row=6, column=0, padx=0, pady=5)

    search_hocphan_button = Button(root, text="Tìm kiếm", command=search_hocphan, width=20)
    search_hocphan_button.grid(row=7, column=0, padx=0, pady=5)

    add_hocphan_button = Button(root, text="Thêm học phần", command=add_hocphan_window, width=20)
    add_hocphan_button.grid(row=12, column=0, padx=0, pady=5)

    conn = connection()
    cur = conn.cursor()
    cur.execute("select * from monhocphan")
    data = cur.fetchall()
    conn.close()

    # Tạo một Treeview widget
    tree = ttk.Treeview(root, columns=(1, 2, 3, 4), show="headings", height=20)

    # Đặt tên các cột
    tree.heading(1, text="Mã học phần")
    tree.heading(2, text="Tên học phần")
    tree.heading(3, text="Số tín chỉ")
    tree.heading(4, text="Mã học kỳ")
    
    # Đặt lại chiều rộng của mã SV, Mã lớp, số tín chỉ.
    tree.column(1, width=310)
    tree.column(2, width=500)
    tree.column(3, width=200)
    tree.column(4, width=200)

    # Hiển thị dữ liệu trong Treeview
    for row in data:
        tree.insert("", "end", values=row)

    # Thêm Treeview vào cửa sổ chính
    tree.grid(row=0, column=1, pady=0, padx=10,rowspan=26)

    # Tạo và thiết lập Scrollbar
    scrollbar = ttk.Scrollbar(root, orient="vertical", command=tree.yview)
    scrollbar.grid(row=0, column=2, sticky="ns",rowspan=26)
    tree.configure(yscrollcommand=scrollbar.set)

    def on_item_click(event):
        # Lấy thông tin của dòng được chọn
        selected_item = tree.selection()[0]
        values = tree.item(selected_item, 'values')

        # Hiển thị cửa sổ chi tiết điểm học phần
        # view_diemhocphan_details(values)
        view_hocphan_details(values)
    # Gán hàm on_item_click khi click vào dòng
    tree.bind('<ButtonRelease-1>', on_item_click)

def view_hocphan_details(hocphan_info):
    # Tạo một cửa sổ mới
    details_window = Toplevel(root)
    details_window.title("SV:" + hocphan_info[1]) 

    # Hiển thị thông tin chi tiết của dòng được chọn
    Label(details_window, text="Mã học phần:").grid(row=0, column=0, padx=10, pady=5)
    mahocphan_entry = Entry(details_window)
    mahocphan_entry.grid(row=0, column=1, padx=10, pady=5)
    mahocphan_entry.insert(0, hocphan_info[0]) 

    Label(details_window, text="Tên học phần:").grid(row=1, column=0, padx=10, pady=5)
    tenhocphan_entry = Entry(details_window)
    tenhocphan_entry.grid(row=1, column=1, padx=10, pady=5)
    tenhocphan_entry.insert(0, hocphan_info[1]) 
    
    Label(details_window, text="Số tín chỉ:").grid(row=2, column=0, padx=10, pady=5)
    sotinchi_entry = Entry(details_window)
    sotinchi_entry.grid(row=2, column=1, padx=10, pady=5)
    sotinchi_entry.insert(0, hocphan_info[2]) 

    Label(details_window, text="Mã học kỳ:").grid(row=3, column=0, padx=10, pady=5)
    mahocky_entry = Entry(details_window)
    mahocky_entry.grid(row=3, column=1, padx=10, pady=5)
    mahocky_entry.insert(0, hocphan_info[3]) 
   
    delete_button = Button(details_window, text="Xoá", command=lambda: delete_hocphan(hocphan_info[0]), fg="white", bg="red",width=20)
    delete_button.grid(row=8, column=0, padx=10, pady=5)

    # #cập nhật tt sinh vien
    def update_hp():
        tenhocphan = tenhocphan_entry.get()
        mahocphan = mahocphan_entry.get()
        sotinchi = sotinchi_entry.get()
        mahocky = mahocky_entry.get()
    
        conn = connection()
        cur = conn.cursor()
        # Cập nhật lại thông tin bảng học phần
        cur.execute("UPDATE monhocphan SET ma_mon=%s, ten_mon=%s, sotinchi=%s, ma_hk=%s WHERE ma_mon=%s",
            (mahocphan, tenhocphan, sotinchi, mahocky, hocphan_info[0]))
        # Cập nhật lại thông tin mã môn của bảng điểm nếu đã tồn tại học phần trước đó.
        cur.execute("UPDATE diemhocphan SET ma_mon=%s WHERE ma_mon=%s", (mahocphan, hocphan_info[0]))
        messagebox.showinfo("Success", "Học phần đã được cập nhật!")
        
        conn.commit()
        conn.close()
        # Đóng cửa sổ chi tiết sau khi cập nhật
        details_window.destroy()
        view_hocphan()
    #xoá học phần 
    def delete_hocphan(ma_hp):
        conn = connection()
        cur = conn.cursor()
        try:
            # Kiểm tra xem học phần tồn tại hay không trước khi xoá
            cur.execute("SELECT ma_mon FROM monhocphan WHERE ma_mon = %s", (ma_hp,))
            existing_sv = cur.fetchone()

            if existing_sv:
                # Nếu học phần tồn tại, thực hiện xoá
                cur.execute("DELETE FROM monhocphan WHERE ma_mon = %s", (ma_hp,))
                conn.commit()
                conn.close()
                messagebox.showinfo("Thông báo", "Xoá học phần thành công")
                details_window.destroy()
                view_hocphan()
            else:
                # Nếu học phần không tồn tại, hiển thị thông báo lỗi
                conn.close()
                messagebox.showerror("Lỗi", "Không tìm thấy học phần có mã này")
        except mysql.connector.Error as e:
            print("Error deleting data from database:", e)
            conn.rollback()
            conn.close()
    update_button = Button(details_window, text="Cập nhật", command=update_hp, fg="black", bg="yellow", width=20)
    update_button.grid(row=8, columnspan=2, column=1, padx=10, pady=5)

def search_hocphan():

    global tree 

    search_query = search_hocphan_entry.get().lower()
    conn = connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM monhocphan WHERE LOWER(ten_mon) LIKE %s OR ma_mon LIKE %s OR ma_hk LIKE %s", ('%' + search_query + '%', '%' + search_query + '%', '%' + search_query + '%'))
    data = cur.fetchall()
    conn.close()

    # Clear the Treeview
    for item in tree.get_children():
        tree.delete(item)

    # Display the filtered data in the Treeview
    for row in data:
        tree.insert("", "end", values=row)
# Hiển thị form thêm mới học phần

def add_hocphan_window():
    add_window = Toplevel(root)
    add_window.title("Thêm Học phần")

    # Hiển thị thông tin chi tiết của dòng được chọn
    Label(add_window, text="Mã học phần:").grid(row=0, column=0, padx=10, pady=5)
    mahocphan_entry = Entry(add_window)
    mahocphan_entry.grid(row=0, column=1, padx=10, pady=5)

    Label(add_window, text="Tên học phần:").grid(row=1, column=0, padx=10, pady=5)
    tenhocphan_entry = Entry(add_window)
    tenhocphan_entry.grid(row=1, column=1, padx=10, pady=5)
    
    Label(add_window, text="Số tín chỉ:").grid(row=2, column=0, padx=10, pady=5)
    sotinchi_entry = Entry(add_window)
    sotinchi_entry.grid(row=2, column=1, padx=10, pady=5)

    Label(add_window, text="Mã học kỳ:").grid(row=3, column=0, padx=10, pady=5)
    mahocky_entry = Entry(add_window)
    mahocky_entry.grid(row=3, column=1, padx=10, pady=5)

    add_button = Button(add_window, text="Thêm Học phần", command=lambda: add_hocphan(
        mahocphan_entry.get(), tenhocphan_entry.get(), sotinchi_entry.get(),
        mahocky_entry.get()
    ), width=20)
    add_button.grid(row=6, column=1, padx=10, pady=10)

def add_hocphan(ma_mon, ten_mon, sotinchi, ma_hk):
    conn = connection()
    cur = conn.cursor()
    try:
        # Kiểm tra xem mã học phần đã tồn tại hay chưa
        cur.execute("SELECT ma_mon FROM monhocphan WHERE ma_mon = %s", (ma_mon,))
        existing_sv = cur.fetchone()

        if existing_sv:
            conn.close()
            messagebox.showerror("Lỗi", "Mã học phần đã tồn tại")
        else:
            # Thêm học phần nếu mã học phần chưa tồn tại
            cur.execute("""
                INSERT INTO monhocphan (ma_mon, ten_mon, sotinchi, ma_hk) 
                VALUES (%s, %s, %s, %s)
            """, (ma_mon, ten_mon, sotinchi, ma_hk))
            conn.commit()
            conn.close()
            messagebox.showinfo("Thông báo", "Thêm học phần thành công")
            search_hocphan()
    except mysql.connector.Error as e:
        print("Error adding data to database:", e)
        conn.rollback()
        conn.close()
# Học phần END #########################################################


#Common Start########################################################
# tổng số tín chỉ đã học 
def total_credit_hours_of_student(ma_sv):
    conn = connection()
    cur = conn.cursor()
    try:
        # Truy vấn cơ sở dữ liệu để lấy tổng số tín chỉ đã học của sinh viên theo mã sinh viên
        cur.execute("""
            SELECT SUM(sotinchi) 
            FROM diemhocphan 
            INNER JOIN monhocphan ON diemhocphan.ma_mon = monhocphan.ma_mon 
            WHERE diemhocphan.ma_sv = %s
        """, (ma_sv,))
        total_credit_hours = cur.fetchone()[0]  # Lấy tổng số tín chỉ từ kết quả truy vấn

        conn.close()

        return total_credit_hours if total_credit_hours else 0  # Trả về tổng số tín chỉ, nếu không có trả về 0
    except mysql.connector.Error as e:
        print("Error retrieving total credit hours:", e)
        conn.close()
        return 0  # Trả về 0 nếu có lỗi khi truy vấn cơ sở dữ liệu


def get_ma_mon_values():
    conn = connection()
    if conn:
        try:
            cur = conn.cursor()
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT ma_mon FROM monhocphan")
            ma_mon_values = [row[0] for row in cur.fetchall()]
            return ma_mon_values
        except mysql.connector.Error as e:
            print("Error fetching ma_mon values:", e)
        finally:
            conn.close()

    return []

def get_ma_lop_values():
    conn = connection()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT ma_lop FROM lop")
            ma_mon_values = [row[0] for row in cur.fetchall()]
            return ma_mon_values
        except mysql.connector.Error as e:
            print("Error fetching ma_mon values:", e)
        finally:
            conn.close()

    return []

def update_sinhvien_combobox(event):
    global ma_sv_combobox
    global ma_lop_combobox
    selected_class = ma_lop_combobox.get()
    students = get_sinhvien_by_lop(selected_class)

    # Xóa toàn bộ mục hiện tại trong combobox sinh viên
    ma_sv_combobox['values'] = ()
    ma_sv_combobox.set("")
    # Thêm danh sách sinh viên mới vào combobox sinh viên
    ma_sv_combobox['values'] = students

def get_sinhvien_by_lop(class_code):
    conn = connection()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("SELECT ma_sv, hoten_sv FROM sinhvien WHERE ma_lop = %s", (class_code,))
            students = cur.fetchall()
            return students
        except mysql.connector.Error as e:
            print("Error fetching students:", e)
        finally:
            conn.close()

    return []

def generate_random_password():
    # Generate a random password with a combination of letters, digits, and punctuation
    password_length = 12  # Adjust the length as needed
    characters = string.ascii_letters + string.digits + string.punctuation
    new_password = ''.join(secrets.choice(characters) for _ in range(password_length))
    return new_password

def clse():
    sys.exit() 
# Common End ########################################################

# Export Excel Start ################################################
def xuat_tatca_sinhvien_diem():
    workbook = Workbook()
    global add_window 
    # Tạo worksheet cho danh sách sinh viên
    ws_sinhvien = workbook.create_sheet(title="Danh sách sinh viên")
    headers_sinhvien = ["Mã SV", "Họ và tên SV", "Ngày sinh", "Giới tính", "Dân tộc", "Nơi sinh", "Mã lớp"]
    ws_sinhvien.append(headers_sinhvien)

    conn = connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM sinhvien")
    data_sinhvien = cur.fetchall()
    conn.close()

    for row in data_sinhvien:
        ws_sinhvien.append(row)

    # Tạo worksheet cho bảng điểm học phần
    ws_diemhocphan = workbook.create_sheet(title="Bảng điểm học phần")
    headers_diemhocphan = [
        "Mã SV", "Họ và tên SV", "Mã lớp", "Mã học phần", "Tên học phần", "Số tín chỉ", "Điểm giữa kỳ", "Điểm cuối kỳ"
    ]
    ws_diemhocphan.append(headers_diemhocphan)

    conn = connection()
    cur = conn.cursor()
    cur.execute("SELECT diemhocphan.ma_sv,hoten_sv,ma_lop,diemhocphan.ma_mon,ten_mon,sotinchi,diem_giua_ky,diem_thi_hp FROM diemhocphan INNER JOIN sinhvien ON diemhocphan.ma_sv=sinhvien.ma_sv INNER JOIN monhocphan ON monhocphan.ma_mon=diemhocphan.ma_mon")
    data_diemhocphan = cur.fetchall()
    conn.close()

    for row in data_diemhocphan:
        ws_diemhocphan.append(row)

    # Lưu file Excel với tên file
    excel_filename = "TatCaSinhVien_DiemHocPhan.xlsx"
    workbook.save(excel_filename)
    print(f"Excel file '{excel_filename}' exported successfully!")
    messagebox.showinfo(f"Export thành công", "File excel đã được export!")

# Export Excel End ################################################

def main_window():
    global login
    # Đóng form login nếu hàm main_window được gọi thành công.
    login.destroy()

    global root
    root = Tk()
    root.title("Quản lý điểm sinh viên")

    t1=Text(root,width=140,height=25)
    t1.grid(row=0,column=1,rowspan=26, padx=10)

    # Đặt màu sắc và font cho các nút button
    button_color = '#3498db'  # Màu xanh dương
    button_font = ('Arial', 12)

    # Danh sách sinh viên
    b1 = Button(root, text="Danh sách sinh viên", command=view_sinhvien, width=40)
    b1.grid(row=0, column=0)
    b1.config(bg=button_color, fg='white', font=button_font)  # Cấu hình màu và font

    # Danh sách học phần
    b3 = Button(root, text="Danh sách học phần", command=view_hocphan, width=40)
    b3.grid(row=2, column=0)
    b3.config(bg=button_color, fg='white', font=button_font)  # Cấu hình màu và font

    # Danh sách điểm học phần
    b4 = Button(root, text="Danh sách điểm học phần", command=view_diemhocphan, width=40)
    b4.grid(row=3, column=0)
    b4.config(bg=button_color, fg='white', font=button_font)  # Cấu hình màu và font

    # Xuất excel
    b8 = Button(root, text="Xuất excel", command=xuat_tatca_sinhvien_diem, width=40)
    b8.grid(row=4, column=0)
    b8.config(bg=button_color, fg='white', font=button_font)  # Cấu hình màu và font

    # Đóng ứng dụng
    b7 = Button(root, text="Đóng ứng dụng", command=clse, width=40)
    b7.grid(row=23, column=0)
    b7.config(bg='red', fg='white', font=button_font)  # Cấu hình màu và font

    root.resizable(False, False)
    root.mainloop()
def login_user():
    username = entry_username.get()
    password = entry_password.get()

    # Add your login logic here (e.g., check username and password)
    # For simplicity, let's assume any non-empty username/password is valid
    conn = connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM dangnhap WHERE username=%s AND password=%s",(username, password))
    user = cur.fetchone()
    conn.close()

    if user:
        main_window()
    else:
        messagebox.showerror("Thông báo", "Tên đăng nhập hoặc mật khẩu không đúng")

def reset_password_window():
    global reset_gmail_entry
    reset_password_window = Toplevel(login)
    reset_password_window.title("Quên mật khẩu")

     # Combobox get mã sinh viên
    Label(reset_password_window, text="Hãy nhập gmail của bạn:").grid(row=0, column=0, padx=20, pady=5)
    reset_gmail_entry = Entry(reset_password_window, width=40)
    reset_gmail_entry.grid(row=1, column=0, padx=20, pady=5)
    reset_password_button = Button(reset_password_window, text="Gửi mật khẩu đến gmail", command=reset_password, width=20, fg="blue")
    reset_password_button.grid(row=3, column=0, columnspan=2, pady=10,padx=20)


def send_reset_email(username, new_password):
    # Replace these values with your email and SMTP server details
    sender_email = "leesintocbien@gmail.com"
    sender_password = "ucfb mezj rpgm miit"
    smtp_server = "smtp.gmail.com"
    smtp_port = 587

    recipient_email = "baoquoc.job@gmail.com"
    subject = "Password Reset"
    body = f"Your new password is: {new_password}"

    # Create the email content
    message = MIMEText(body)
    message["Subject"] = subject
    message["From"] = sender_email
    message["To"] = recipient_email

    # Connect to the SMTP server and send the email
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, [recipient_email], message.as_string())

def reset_password():
    email = reset_gmail_entry.get()
    # Random password mới để gửi đến mail
    new_password = generate_random_password()

    try:
        # Update the password in the database
        conn = connection()
        cur = conn.cursor()
        cur.execute("select emai from dangnhap where emai=%s", (email,))
        email_result = cur.fetchone()
        conn.commit()

       # Cập nhật lại mật khẩu người quản lý trong hệ thống.
        if email_result :
            conn = connection()
            cur = conn.cursor()
            cur.execute("UPDATE dangnhap SET password=%s WHERE emai=%s", (new_password, email))
            conn.commit()
            conn.close()

            # Send the new password via email
            send_reset_email(email, new_password)

            messagebox.showinfo("Thông báo", "Mật khẩu mới đã được gửi đến email của bạn.")
        else :
            messagebox.showinfo("Thông báo", "Có vẻ như thông tin đăng nhập của bạn không tồn tại trong hệ thống, hãy liên hệ bộ phận IT nhé.")
    except Exception as e:
        messagebox.showerror("Lỗi", f"Có lỗi xảy ra: {str(e)}")
    finally:
        conn.close()
    

if __name__ == "__main__":
    login = Tk()
    login.title("Đăng nhập vào hệ thống quản lý điểm")

    label_username = Label(login, text="Tên đăng nhập:")
    label_username.grid(row=0, column=0, padx=30, pady=40)
    entry_username = Entry(login, width=40)
    entry_username.grid(row=0, column=1, padx=30, pady=10)

    label_password = Label(login, text="Mật khẩu:")
    label_password.grid(row=1, column=0, padx=30, pady=10)
    entry_password = Entry(login, show="*", width=40)
    entry_password.grid(row=1, column=1, padx=30, pady=10)

    b1 = Button(login, text="Đăng nhập", command=login_user, width=40, fg="green")
    b1.grid(row=2, column=0, columnspan=2, pady=40)
    # Add a "Forgot Password" button
    b2 = Button(login, text="Quên mật khẩu", command=reset_password_window, width=40, fg="blue")
    b2.grid(row=3, column=0, columnspan=2, pady=10)

    login.resizable(False, False)
    login.mainloop()
